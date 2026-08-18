# Part of Odoo. See LICENSE file for full copyright and licensing details.
import base64
import csv
import io

from odoo import api, fields, models
from odoo.exceptions import UserError

SOURCE_SYSTEM = 'google_sheets'
# Types de personnes contre lesquels l'export etudiants se rapproche. Un
# matricule d'employe qui remonterait dans cet export est un conflit, pas un
# rapprochement : cf. his.person._find_or_flag_match.
CANDIDATE_TYPES = ('etudiant', 'candidat')

OUTCOME_SELECTION = [
    ('created', "Fiche creee"),
    ('updated', "Fiche mise a jour"),
    ('flagged', "A arbitrer"),
    ('confirmed', "Rapprochement confirme"),
    ('rejected', "Rapprochement refuse, fiche creee"),
    ('conflict', "Conflit - ligne rejetee"),
]

ROW_FIELDS = (
    'matricule_institutionnel', 'nom_latin', 'nom_arabe',
    'email_institutionnel', 'email_personnel', 'telephone', 'external_ref',
)

# Plusieurs en-tetes possibles pour une meme colonne : l'export est produit a
# la main dans une feuille, ses intitules varient d'une extraction a l'autre.
COLUMN_ALIASES = {
    'matricule_institutionnel': (
        'matricule', 'matricule institutionnel', 'id institutionnel',
    ),
    'nom_latin': ('nom latin', 'nom', 'name', 'nom complet'),
    'nom_arabe': ('nom arabe', 'nom ar', 'arabe'),
    'email_institutionnel': (
        'email institutionnel', 'email pro', 'mail institutionnel',
    ),
    'email_personnel': ('email personnel', 'email', 'mail', 'e-mail'),
    'telephone': ('telephone', 'tel', 'phone', 'gsm', 'mobile'),
    'external_ref': ('external ref', 'reference', 'ref', 'id', 'identifiant', 'row id'),
}


def _normalize_header(header):
    return ' '.join((header or '').strip().lower().replace('_', ' ').split())


class HisPersonImport(models.TransientModel):
    _name = 'his.person.import'
    _description = "Import des personnes depuis l'export Google Sheets"

    file = fields.Binary(string="Fichier export (CSV ou XLSX)", required=True)
    filename = fields.Char(string="Nom du fichier")
    state = fields.Selection(
        selection=[('upload', "Depot"), ('review', "Resultat")],
        default='upload', string="Etape",
    )
    line_ids = fields.One2many('his.person.import.line', 'import_id', string="Lignes")

    created_count = fields.Integer(string="Creees", compute='_compute_counts')
    updated_count = fields.Integer(string="Mises a jour", compute='_compute_counts')
    flagged_count = fields.Integer(string="A arbitrer", compute='_compute_counts')
    conflict_count = fields.Integer(string="Conflits", compute='_compute_counts')

    @api.depends('line_ids.outcome')
    def _compute_counts(self):
        for wizard in self:
            outcomes = wizard.line_ids.mapped('outcome')
            wizard.created_count = outcomes.count('created') + outcomes.count('rejected')
            wizard.updated_count = outcomes.count('updated') + outcomes.count('confirmed')
            wizard.flagged_count = outcomes.count('flagged')
            wizard.conflict_count = outcomes.count('conflict')

    # --- Lecture du fichier -------------------------------------------------

    def _read_rows(self):
        """Retourne une liste de dicts {champ his.person: valeur}, un par ligne."""
        content = base64.b64decode(self.file)
        name = (self.filename or '').lower()
        if name.endswith(('.xlsx', '.xlsm')):
            raw_rows = self._read_xlsx(content)
        else:
            raw_rows = self._read_csv(content)
        rows = self._map_rows(raw_rows)
        if not rows:
            raise UserError("Le fichier ne contient aucune ligne exploitable.")
        return rows

    def _read_csv(self, content):
        for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UserError("Encodage du fichier CSV non reconnu.")
        # Sniff : la feuille est exportee tantot en virgule, tantot en
        # point-virgule selon la locale du poste qui l'exporte.
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=',;\t').delimiter
        except csv.Error:
            delimiter = ','
        return list(csv.reader(io.StringIO(text), delimiter=delimiter))

    def _read_xlsx(self, content):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(
                "La lecture XLSX necessite openpyxl, absent de cet environnement. "
                "Exportez la feuille en CSV et relancez l'import."
            )
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [
            ['' if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]

    def _map_rows(self, raw_rows):
        raw_rows = [row for row in raw_rows if any((cell or '').strip() for cell in row)]
        if not raw_rows:
            return []
        headers = [_normalize_header(cell) for cell in raw_rows[0]]
        mapping = {}
        for field, aliases in COLUMN_ALIASES.items():
            normalized = {_normalize_header(alias) for alias in aliases}
            normalized.add(_normalize_header(field))
            for index, header in enumerate(headers):
                if header in normalized and field not in mapping:
                    mapping[field] = index
        if 'nom_latin' not in mapping:
            raise UserError(
                "Aucune colonne de nom reconnue dans l'en-tete du fichier. "
                "Intitules acceptes : %s." % ", ".join(COLUMN_ALIASES['nom_latin'])
            )
        rows = []
        for position, raw in enumerate(raw_rows[1:], start=2):
            row = {
                field: (raw[index].strip() if index < len(raw) and raw[index] else '')
                for field, index in mapping.items()
            }
            if not row.get('nom_latin'):
                continue  # ligne vide ou ligne de separation dans la feuille
            # Sans reference propre dans la source, la position sert de cle de
            # rejeu : reimporter le meme fichier doit retomber sur les memes
            # fiches, pas en creer de nouvelles.
            if not row.get('external_ref'):
                row['external_ref'] = '%s#L%s' % (self.filename or 'export', position)
            row['source_system'] = SOURCE_SYSTEM
            rows.append(row)
        return rows

    # --- Traitement ---------------------------------------------------------

    def action_import(self):
        self.ensure_one()
        Person = self.env['his.person']
        Line = self.env['his.person.import.line']
        self.line_ids.unlink()

        for row in self._read_rows():
            match = Person._find_or_flag_match(row, types=CANDIDATE_TYPES)
            line_vals = dict(
                {field: row.get(field) or False for field in ROW_FIELDS},
                import_id=self.id,
                score=match['score'],
                candidate_person_id=match['person'].id if match['person'] else False,
            )

            if match['conflict']:
                # Arret net sur la ligne. Ni fusion, ni ecrasement : deux
                # personnes differentes derriere un meme matricule est un
                # probleme de donnee source, pas un cas a trancher ici.
                line_vals.update(outcome='conflict', message=match['conflict'])
            elif match['method'] == 'deterministic' and match['person']:
                self._update_person(match['person'], row)
                line_vals.update(
                    outcome='updated', message="Rapprochement deterministe.",
                    person_id=match['person'].id,
                )
            elif match['method'] == 'probabilistic':
                # Jamais de lien automatique, quel que soit le score.
                line_vals.update(
                    outcome='flagged',
                    message="Correspondance probable (%d%%) : a confirmer ou refuser."
                            % round(match['score'] * 100),
                )
            else:
                method = 'deterministic' if row.get('matricule_institutionnel') else 'new'
                person = self._create_person(row, method)
                line_vals.update(outcome='created', person_id=person.id)

            self._log_line(Line.create(line_vals))

        self.state = 'review'
        return self._reopen()

    @api.model
    def _person_vals(self, row):
        return {
            field: row.get(field) or False
            for field in (
                'nom_latin', 'nom_arabe', 'email_institutionnel',
                'email_personnel', 'telephone', 'external_ref',
            )
        }

    @api.model
    def _create_person(self, row, match_method):
        vals = dict(
            self._person_vals(row),
            type_personne='etudiant',
            source_system=SOURCE_SYSTEM,
            match_method=match_method,
        )
        if row.get('matricule_institutionnel'):
            # Valeur deja attribuee cote source : elle est reprise telle
            # quelle, his_person_core n'en emet pas de nouvelle.
            vals['matricule_institutionnel'] = row['matricule_institutionnel']
        return self.env['his.person'].sudo().create(vals)

    @api.model
    def _update_person(self, person, row):
        """Met a jour les champs non identitaires. Le matricule n'est jamais touche."""
        vals = {
            field: value
            for field, value in self._person_vals(row).items()
            if value and value != person[field]
        }
        if vals:
            person.sudo().write(vals)
        return person

    def _log_line(self, lines):
        """Trace persistante de chaque decision, y compris celles sans fiche.

        Un modele dedie plutot que le seul chatter : un conflit ou un refus
        n'est rattache a aucune fiche, un message de chatter n'aurait nulle
        part ou se poser. Les fiches reellement touchees recoivent en plus
        leur message de chatter.
        """
        Log = self.env['his.person.sync.log'].sudo()
        for line in lines:
            Log.create({
                'person_id': line.person_id.id or line.candidate_person_id.id or False,
                'source_system': SOURCE_SYSTEM,
                'external_ref': line.external_ref,
                'nom_source': line.nom_latin,
                'matricule_source': line.matricule_institutionnel,
                'outcome': line.outcome,
                'score': line.score,
                'message': line.message,
            })
            if line.person_id:
                line.person_id.sudo().message_post(
                    body="Import Google Sheets (%s) : %s" % (
                        line.external_ref or '-', line.message or line.outcome,
                    ),
                )

    def _reopen(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class HisPersonImportLine(models.TransientModel):
    _name = 'his.person.import.line'
    _description = "Ligne d'import de personnes"
    _order = 'outcome, id'

    import_id = fields.Many2one('his.person.import', required=True, ondelete='cascade')

    matricule_institutionnel = fields.Char(string="Matricule (source)", readonly=True)
    nom_latin = fields.Char(string="Nom (latin)", readonly=True)
    nom_arabe = fields.Char(string="Nom (arabe)", readonly=True)
    email_institutionnel = fields.Char(string="Email institutionnel", readonly=True)
    email_personnel = fields.Char(string="Email personnel", readonly=True)
    telephone = fields.Char(string="Telephone", readonly=True)
    external_ref = fields.Char(string="Reference source", readonly=True)

    outcome = fields.Selection(selection=OUTCOME_SELECTION, string="Resultat", readonly=True)
    message = fields.Char(string="Detail", readonly=True)
    score = fields.Float(string="Score", digits=(3, 2), readonly=True)
    person_id = fields.Many2one('his.person', string="Fiche resultante", readonly=True)
    candidate_person_id = fields.Many2one(
        'his.person', string="Fiche proposee", readonly=True,
        help="Meilleure correspondance trouvee. Tant qu'elle n'est pas confirmee, "
             "rien n'est rattache.",
    )

    def action_confirm_match(self):
        """L'administrateur reconnait la personne : on rattache, on trace qui et quand."""
        for line in self:
            if line.outcome != 'flagged' or not line.candidate_person_id:
                continue
            person = line.candidate_person_id
            line.import_id._update_person(person, line._to_row())
            person.sudo().write({
                'match_method': 'probabilistic',
                'matched_by': self.env.user.id,
                'matched_on': fields.Datetime.now(),
            })
            line.write({
                'outcome': 'confirmed', 'person_id': person.id,
                'message': "Rapprochement (%d%%) confirme par %s." % (
                    round(line.score * 100), self.env.user.display_name,
                ),
            })
            line.import_id._log_line(line)
        return self[:1].import_id._reopen()

    def action_reject_match(self):
        """Ce n'est pas la meme personne : nouvelle fiche, nouveau matricule."""
        for line in self:
            if line.outcome != 'flagged':
                continue
            person = line.import_id._create_person(line._to_row(), 'new')
            line.write({
                'outcome': 'rejected', 'person_id': person.id,
                'message': "Rapprochement refuse par %s : fiche distincte creee."
                           % self.env.user.display_name,
            })
            line.import_id._log_line(line)
        return self[:1].import_id._reopen()

    def _to_row(self):
        self.ensure_one()
        return {field: self[field] or '' for field in ROW_FIELDS}
